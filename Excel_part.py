import openpyxl as xl

wb = xl.load_workbook('loop_diagrams.xlsx')

sheet = wb['check_sheet']

list_of_loops = []
for row in range(2, sheet.max_row + 1):
    if sheet.cell(row, 1).value is None:
        break
    loop = {"DocNumber": sheet.cell(row, 1).value, "Name": sheet.cell(row, 2).value, "Link": sheet.cell(row, 3).value}
    print(loop)
    list_of_loops.append(loop)

wb.close()