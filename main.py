import openpyxl as xl
import os
from autoRLMU import AnnotationMakerOld, AnnotationMakerNew


os.system("cls")
wb = xl.load_workbook('loop_diagrams.xlsx')

try:
    sheet = wb['check_sheet']

    # for each row in 'check_sheet' sheet
    for row in range(2, sheet.max_row + 1):
        # if an empty row - stop
        if sheet.cell(row, 1).value is None:
            break

        if sheet.cell(row, 3).value not in (None, ""):
            continue

        # clear result
        sheet.cell(row, 4).value = ''

        # save in dict
        # loop = {"Doc Number": sheet.cell(row, 1).value, "Link": sheet.cell(row, 2).hyperlink.target}
        loop = {"Doc Number": sheet.cell(row, 1).value, "Link": sheet.cell(row, 2).value}

        # creating object
        pdf_path_annotated: str = f'pdfs/{loop["Doc Number"]}.pdf'
        loop_drawing = AnnotationMakerOld()

        # trying to make a redline, comment the cropping if no need
        # standard crop new for: x0=730 y0=12 w=446 h=698
        # crop for small new:
        # loop_drawing.set_crop_rectangle_wh(950, 30, 215, 650)
        # standard crop for old: x0=989 y0=62 w=163 h=562
        # loop_drawing.set_crop_rectangle_wh(898, 30, 215, 650)
        is_redlined_successfully = loop_drawing.make_redline(loop['Link'], dpi=300)
        if is_redlined_successfully:
            loop["Result"] = 'Success'
        else:
            loop["Result"] = loop_drawing.get_error_description()

        # filling the result in the Excel file
        sheet.cell(row, 3).value = loop["Result"]
        sheet.cell(row, 4).value = loop_drawing.get_log()

        try:
            wb.save('loop_diagrams.xlsx')
        except Exception as e:
            print(f'Cannot save the excel file: {str(e)}')
finally:
    wb.close()
